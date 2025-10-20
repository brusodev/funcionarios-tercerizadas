import pandas as pd
import openpyxl

# Abre o arquivo com openpyxl para ver formatação
wb = openpyxl.load_workbook('analise_detalhada.xlsx')
ws = wb['743']

print("="*80)
print("ANÁLISE COMPLETA DO EXCEL - PDF 0743")
print("="*80)

# Verifica TODAS as colunas
print("\nCOLUNAS PRESENTES:")
for col_idx in range(1, ws.max_column + 1):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    header = ws[f'{col_letter}1'].value
    print(f"   Coluna {col_letter}: {header}")

# Conta linhas com valores em cada coluna
print(f"\nCONTAGEM DE VALORES POR COLUNA:")
for col_idx in range(1, ws.max_column + 1):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    
    count = 0
    sum_values = 0
    for row in range(2, ws.max_row + 1):
        cell = ws[f'{col_letter}{row}']
        if cell.value is not None and cell.value != '':
            count += 1
            # Tenta somar se for número
            try:
                val = float(str(cell.value).replace(',', '.'))
                sum_values += val
            except:
                pass
    
    print(f"   Coluna {col_letter}: {count} células com valor")
    if sum_values > 0:
        print(f"      Soma numérica: {sum_values}")

# Agora vê se há células destacadas/coloridas
print(f"\n{'='*80}")
print("VERIFICANDO SE HÁ CÉLULAS COLORIDAS/DESTACADAS:")
print(f"{'='*80}")

from openpyxl.styles import PatternFill
from collections import defaultdict

cores = defaultdict(list)
linhas_destacadas = set()

for row in range(2, min(100, ws.max_row + 1)):  # Primeiras 100 linhas
    for col_idx in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cell = ws[f'{col_letter}{row}']
        
        # Verifica cor de fundo
        if cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000':
            cor = cell.fill.start_color.index
            if cor not in ['00000000', 'FFFFFFFF', None]:  # Ignora preto e branco
                cores[cor].append(row)
                linhas_destacadas.add(row)

if cores:
    print(f"\nEncontradas {len(linhas_destacadas)} linhas com destaque:")
    for cor, rows in cores.items():
        print(f"   Cor {cor}: {len(rows)} linhas")
else:
    print("\nNenhuma célula colorida encontrada nas primeiras 100 linhas.")

# Lê o DataFrame normalmente para análise
df = pd.read_excel('analise_detalhada.xlsx', sheet_name='743')

print(f"\n{'='*80}")
print("ESTRUTURA DO DATAFRAME:")
print(f"{'='*80}")
print(f"Total de linhas: {len(df)}")
print(f"Total de colunas: {len(df.columns)}")
print(f"\nPrimeiras colunas:")
for i, col in enumerate(df.columns[:10]):
    print(f"   Coluna {i}: {col}")

# Mostra primeiras 5 linhas completas
print(f"\n{'='*80}")
print("PRIMEIRAS 5 LINHAS DO EXCEL:")
print(f"{'='*80}")
for idx in range(min(5, len(df))):
    row = df.iloc[idx]
    print(f"\nLinha {idx+2}:")
    for i, val in enumerate(row):
        if pd.notna(val) and val != '':
            print(f"   Coluna {i} ({chr(65+i)}): {val}")

wb.close()
