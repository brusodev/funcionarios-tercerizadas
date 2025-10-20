import openpyxl
import PyPDF2
import re

# Ler totais do Excel
wb = openpyxl.load_workbook('ADM_TOTAL.xlsx')

excel_totais = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    total = 0
    
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        n = row[0]
        if n and isinstance(n, (int, float)):
            total += int(n) if isinstance(n, int) else int(n)
    
    excel_totais[sheet_name] = total

# Contagem do script (original, com regex \d+,\d+)
pdfs_lista = [
    ('0733', 'ADM 0800100_0733_2025 de 30_09_2025.PDF'),
    ('0734', 'ADM 0800100_0734_2025 de 30_09_2025.PDF'),
    ('0735', 'ADM 0800100_0735_2025 de 30_09_2025.PDF'),
    ('0736', 'ADM 0800100_0736_2025 de 30_09_2025.PDF'),
    ('0737', 'ADM 0800100_0737_2025 de 30_09_2025.PDF'),
    ('0738', 'ADM 0800100_0738_2025 de 30_09_2025.PDF'),
    ('0739', 'ADM 0800100_0739_2025 de 30_09_2025.PDF'),
    ('0740', 'ADM 0800100_0740_2025 de 30_09_2025.PDF'),
    ('0741', 'ADM 0800100_0741_2025 de 30_09_2025.PDF'),
    ('0742', 'ADM 0800100_0742_2025 de 30_09_2025.PDF'),
    ('0743', 'ADM 0800100_0743_2025 de 30_09_2025.PDF'),
    ('0744', 'ADM 0800100_0744_2025 de 30_09_2025.PDF'),
    ('0745', 'ADM 0800100_0745_2025 de 30_09_2025.PDF'),
    ('0746', 'ADM 0800100_0746_2025 de 30_09_2025.PDF'),
    ('0747', 'ADM 0800100_0747_2025 de 30092025.PDF'),
]

script_totais = {}
linhas_faltantes = {}

for pdf_num, pdf_nome in pdfs_lista:
    total = 0
    linhas_encontradas = 0
    
    try:
        with open(pdf_nome, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            text = ''
            for page in reader.pages:
                text += page.extract_text()
        
        lines = text.split('\n')
        
        for linha in lines:
            # Padrão 1: números simples com vírgula (4,00 un)
            match = re.match(r'^(\d+,\d+)\s+un\s+', linha)
            if match:
                qty_str = match.group(1).replace(',', '.')
                qtd = int(float(qty_str))
                total += qtd
                linhas_encontradas += 1
            else:
                # Padrão 2: números com ponto de milhar (1.185,00 un)
                match2 = re.match(r'^(\d+\.\d+,\d+)\s+un\s+', linha)
                if match2:
                    qty_str = match2.group(1).replace('.', '').replace(',', '.')
                    qtd = int(float(qty_str))
                    total += qtd
                    linhas_encontradas += 1
                else:
                    # Padrão 3: kg com vírgula (0,49 kg)
                    match3 = re.match(r'^(\d+,\d+)\s+kg\s+', linha)
                    if match3:
                        qty_str = match3.group(1).replace(',', '.')
                        qtd_kg = float(qty_str)
                        qtd = int(qtd_kg / 0.5)
                        total += qtd
                        linhas_encontradas += 1
                    else:
                        # Padrão 4: kg com ponto de milhar (1.185,00 kg)
                        match4 = re.match(r'^(\d+\.\d+,\d+)\s+kg\s+', linha)
                        if match4:
                            qty_str = match4.group(1).replace('.', '').replace(',', '.')
                            qtd_kg = float(qty_str)
                            qtd = int(qtd_kg / 0.5)
                            total += qtd
                            linhas_encontradas += 1
        
        script_totais[pdf_num] = total
        linhas_faltantes[pdf_num] = linhas_encontradas
        
    except Exception as e:
        print(f"Erro ao processar {pdf_nome}: {e}")
        script_totais[pdf_num] = 0

# Comparação
print("="*140)
print("COMPARACAO: EXCEL (Sua Contagem) vs SCRIPT (Original)")
print("="*140 + "\n")

print(f"{'PDF':<6} {'Excel':<12} {'Script':<12} {'Diferenca':<15} {'Linhas':<8} {'Status':<20}\n")

total_excel = 0
total_script = 0
erros_encontrados = []

for pdf_num in sorted(excel_totais.keys()):
    excel = excel_totais[pdf_num]
    script = script_totais.get(pdf_num, 0)
    diferenca = script - excel
    linhas = linhas_faltantes.get(pdf_num, 0)
    
    total_excel += excel
    total_script += script
    
    if diferenca == 0:
        status = "OK"
    elif diferenca > 0:
        status = f"Script +{diferenca}"
        erros_encontrados.append((pdf_num, excel, script, diferenca))
    else:
        status = f"Faltam {-diferenca}"
        erros_encontrados.append((pdf_num, excel, script, diferenca))
    
    print(f"{pdf_num:<6} {excel:<12,d} {script:<12,d} {diferenca:+d}             {linhas:<8} {status:<20}")

print(f"\n{'─'*140}")
print(f"{'TOTAL':<6} {total_excel:<12,d} {total_script:<12,d} {total_script - total_excel:+d}")
print("="*140 + "\n")

if erros_encontrados:
    print("ERROS ENCONTRADOS:\n")
    for pdf_num, excel, script, dif in erros_encontrados:
        print(f"PDF {pdf_num}: Excel={excel:,}, Script={script:,}, Diferenca={dif:+,}")
else:
    print("SEM ERROS - TUDO BATE!")

print()
