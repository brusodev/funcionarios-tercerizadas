import openpyxl
import json
import re
from collections import defaultdict

def extrair_modelo_descricao(descricao):
    """
    Extrai o modelo do celular da descrição
    Ex: "3,00 un TELEFONE CELULAR BLU BOLD Z4 MUSIC/// R$ 348,00"
    Retorna: "BLU BOLD Z4 MUSIC"
    """
    if not descricao:
        return None
    
    descricao = str(descricao).strip()
    
    # Remove quantidade no início (ex: "3,00 un")
    descricao = re.sub(r'^\s*\d+[\.,]\d+\s+un\s+', '', descricao, flags=re.IGNORECASE)
    # Remove valor no final (ex: "R$ 348,00")
    descricao = re.sub(r'\s*R\$\s*[\d.,]+\s*$', '', descricao)
    # Remove marcadores extras (///, ----, etc)
    descricao = re.sub(r'[/\-#]+$', '', descricao)
    # Remove "SN OU IMEI NO" e similares
    descricao = re.sub(r'[,\s]+(SN OU IMEI NO|SERIE|IMEI|SN)', '', descricao, flags=re.IGNORECASE)
    # Remove parênteses e conteúdo
    descricao = re.sub(r'\([^)]*\)', '', descricao)
    
    descricao = descricao.strip()
    
    return descricao if descricao else "DESCONHECIDO"

def extrair_quantidade(descricao):
    """
    Extrai a quantidade da descrição
    Ex: "3,00 un ..." retorna 3
    """
    if not descricao:
        return 0
    
    match = re.match(r'^\s*(\d+[\.,]\d+)\s+un', str(descricao), flags=re.IGNORECASE)
    if match:
        qtd_str = match.group(1).replace(',', '.')
        return int(float(qtd_str))
    
    return 0

def processar_excel():
    """
    Processa o Excel e cria estrutura de modelos por ADM
    """
    
    wb = openpyxl.load_workbook('ADM_TOTAL.xlsx')
    
    # Estrutura final: { "ADM": { "modelo": { "quantidade": X, "pdfs": ["0733", "0734"] } } }
    resultado = {
        'data_geracao': '2025-10-17',
        'total_geral': {
            'aparelhos': 0,
            'modelos_unicos': 0,
            'valor': 'R$ 30.241.453,50'
        },
        'adms': {}
    }
    
    adm_totais = defaultdict(int)  # Para validar total por ADM
    modelos_globais = defaultdict(int)  # Para contar modelos únicos
    
    print("=" * 80)
    print("PROCESSANDO EXCEL ADM_TOTAL.xlsx")
    print("=" * 80)
    print()
    
    for sheet_name in sorted(wb.sheetnames):
        print(f"Processando ADM {sheet_name}...", end=" ", flush=True)
        
        ws = wb[sheet_name]
        
        adm_dados = {
            'total_aparelhos': 0,
            'modelos': {}
        }
        
        linha_contador = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]:
                continue
            
            quantidade = extrair_quantidade(str(row[1]) if row[1] else "")
            if quantidade == 0:
                quantidade = row[0] if isinstance(row[0], int) else 0
            
            modelo = extrair_modelo_descricao(row[1])
            
            if quantidade > 0 and modelo:
                linha_contador += 1
                adm_dados['total_aparelhos'] += quantidade
                adm_totais[sheet_name] += quantidade
                
                # Agrupa modelos iguais
                if modelo not in adm_dados['modelos']:
                    adm_dados['modelos'][modelo] = {
                        'quantidade': 0,
                        'linhas': 0
                    }
                
                adm_dados['modelos'][modelo]['quantidade'] += quantidade
                adm_dados['modelos'][modelo]['linhas'] += 1
                
                # Conta modelo globalmente
                modelos_globais[modelo] += quantidade
        
        # Ordena modelos por quantidade decrescente
        modelos_ordenados = dict(sorted(
            adm_dados['modelos'].items(),
            key=lambda x: x[1]['quantidade'],
            reverse=True
        ))
        
        adm_dados['modelos'] = modelos_ordenados
        adm_dados['total_modelos_unicos'] = len(modelos_ordenados)
        
        resultado['adms'][sheet_name] = adm_dados
        
        print(f"✓ {adm_dados['total_aparelhos']:,} aparelhos | {len(modelos_ordenados)} modelos únicos")
    
    # Calcula totais gerais
    resultado['total_geral']['aparelhos'] = sum(adm_totais.values())
    resultado['total_geral']['modelos_unicos'] = len(modelos_globais)
    resultado['total_geral']['adms'] = len(resultado['adms'])
    
    # Adiciona ranking de modelos mais comuns
    modelos_ranking = sorted(
        modelos_globais.items(),
        key=lambda x: x[1],
        reverse=True
    )
    
    resultado['top_20_modelos'] = [
        {'modelo': modelo, 'quantidade': qtd}
        for modelo, qtd in modelos_ranking[:20]
    ]
    
    return resultado, adm_totais

def main():
    resultado, adm_totais = processar_excel()
    
    print()
    print("=" * 80)
    print("RESUMO GERAL")
    print("=" * 80)
    print()
    print(f"Total de Aparelhos: {resultado['total_geral']['aparelhos']:,}")
    print(f"Modelos Únicos: {resultado['total_geral']['modelos_unicos']}")
    print(f"ADMs Processados: {resultado['total_geral']['adms']}")
    print()
    
    print("=" * 80)
    print("TOP 20 MODELOS MAIS COMUNS")
    print("=" * 80)
    print()
    
    for i, item in enumerate(resultado['top_20_modelos'], 1):
        print(f"{i:2}. {item['modelo']:60} | {item['quantidade']:6,} aparelhos")
    
    print()
    print("=" * 80)
    print("VALIDAÇÃO POR ADM")
    print("=" * 80)
    print()
    
    total_validacao = 0
    for adm in sorted(adm_totais.keys()):
        qtd = adm_totais[adm]
        total_validacao += qtd
        print(f"ADM {adm}: {qtd:,} aparelhos ✓")
    
    print()
    print(f"Total Geral Validado: {total_validacao:,} aparelhos")
    print(f"Total Esperado: 33.215 aparelhos")
    
    if total_validacao == 33215:
        print("✅ VALIDAÇÃO PERFEITA!")
    else:
        print(f"⚠️  DIFERENÇA: {total_validacao - 33215:+d} aparelhos")
    
    print()
    print("=" * 80)
    print("SALVANDO ARQUIVO JSON...")
    print("=" * 80)
    print()
    
    # Salva arquivo JSON
    with open('MEGA_JSON_MODELOS_COMPLETO.json', 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)
    
    print("✓ Arquivo salvo: MEGA_JSON_MODELOS_COMPLETO.json")
    print()
    
    # Também salva um JSON simplificado com apenas os modelos e totais
    json_simplificado = {
        'total_aparelhos': resultado['total_geral']['aparelhos'],
        'modelos_unicos': resultado['total_geral']['modelos_unicos'],
        'valor_total': 'R$ 30.241.453,50',
        'modelos': {}
    }
    
    # Monta lista global de modelos
    for adm, dados in resultado['adms'].items():
        for modelo, info in dados['modelos'].items():
            if modelo not in json_simplificado['modelos']:
                json_simplificado['modelos'][modelo] = {
                    'quantidade_total': 0,
                    'adms': {}
                }
            
            json_simplificado['modelos'][modelo]['quantidade_total'] += info['quantidade']
            json_simplificado['modelos'][modelo]['adms'][adm] = info['quantidade']
    
    # Ordena modelos por quantidade
    json_simplificado['modelos'] = dict(sorted(
        json_simplificado['modelos'].items(),
        key=lambda x: x[1]['quantidade_total'],
        reverse=True
    ))
    
    with open('MEGA_JSON_MODELOS_SIMPLIFICADO.json', 'w', encoding='utf-8') as f:
        json.dump(json_simplificado, f, ensure_ascii=False, indent=2)
    
    print("✓ Arquivo salvo: MEGA_JSON_MODELOS_SIMPLIFICADO.json")
    print()
    print("✅ PROCESSAMENTO CONCLUÍDO COM SUCESSO!")

if __name__ == '__main__':
    main()
